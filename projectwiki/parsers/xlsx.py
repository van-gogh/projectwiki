from __future__ import annotations

from pathlib import Path

from ..models import ParsedBlock


def parse_xlsx(path: Path) -> list[ParsedBlock]:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to parse .xlsx files") from exc

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    blocks: list[ParsedBlock] = []
    for sheet in wb.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(v).strip() if v is not None else f"col_{i+1}" for i, v in enumerate(rows[0])]
        for row_idx, row in enumerate(rows[1:], start=2):
            values = {headers[i] if i < len(headers) else f"col_{i+1}": row[i] for i in range(len(row))}
            text = "; ".join(f"{k}: {v}" for k, v in values.items() if v not in (None, ""))
            if text:
                blocks.append(
                    ParsedBlock(
                        "table_row",
                        text,
                        {"sheet": sheet.title, "row": row_idx, "range": f"A{row_idx}:{sheet.max_column}{row_idx}"},
                        {"headers": headers, "values": values},
                    )
                )
    return blocks
